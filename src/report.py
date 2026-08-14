"""Result export.

Everything the study produces lands in two places at once: a flat CSV per table
in ``outputs/tables`` (diff-able, loadable, what a co-author will actually reuse)
and one multi-sheet ``.xlsx`` workbook that mirrors the same tables with
formatting, so the whole result set travels as a single file.
"""
from __future__ import annotations

import os
import re
from typing import Dict, List, Optional, Sequence

import numpy as np
import pandas as pd

from .config import DIR_FIGURES, DIR_TABLES, OUT_ROOT

_REGISTRY: "dict[str, pd.DataFrame]" = {}
_CAPTIONS: "dict[str, str]" = {}


def register(name: str, df: pd.DataFrame, caption: str = "") -> pd.DataFrame:
    """Save one table as CSV and queue it for the Excel workbook."""
    if df is None or (hasattr(df, "empty") and df.empty):
        return df
    out = df.copy()
    out.to_csv(os.path.join(DIR_TABLES, f"{name}.csv"), index=False)
    _REGISTRY[name] = out
    if caption:
        _CAPTIONS[name] = caption
    return out


def registered() -> Dict[str, pd.DataFrame]:
    return dict(_REGISTRY)


def adopt_existing_tables() -> int:
    """Pull in tables written by earlier runs of the pipeline.

    The study is split across invocations (main comparison, then ablation), so the
    workbook has to gather every CSV in ``outputs/tables`` — not just the ones
    this process happened to produce. Tables registered in-process win, since they
    are the fresher version of the same name.
    """
    added = 0
    for fn in sorted(os.listdir(DIR_TABLES)):
        if not fn.endswith(".csv"):
            continue
        name = fn[:-4]
        if name in _REGISTRY:
            continue
        try:
            _REGISTRY[name] = pd.read_csv(os.path.join(DIR_TABLES, fn))
            added += 1
        except Exception:
            continue
    ordered = dict(sorted(_REGISTRY.items(), key=lambda kv: (not kv[0][:1].isupper(), kv[0])))
    _REGISTRY.clear()
    _REGISTRY.update(ordered)
    return added


def _sheet_name(name: str) -> str:
    """Excel sheet names: <=31 chars, no []:*?/\\ ."""
    s = re.sub(r"[\[\]:*?/\\]", "-", name)
    return s[:31]


def round_numeric(df: pd.DataFrame, digits: int = 4) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if pd.api.types.is_float_dtype(out[c]):
            out[c] = out[c].round(digits)
    return out


def format_mean_ci(point: float, lo: float, hi: float, digits: int = 3) -> str:
    if any(not np.isfinite(v) for v in (point, lo, hi)):
        return "n/a"
    return f"{point:.{digits}f} [{lo:.{digits}f}, {hi:.{digits}f}]"


def paper_table(summary: pd.DataFrame, ci: Optional[pd.DataFrame] = None,
                metrics: Sequence[str] = ("Dice_agg", "Dice", "IoU", "Precision",
                                          "Recall", "HD95", "ASSD", "NSD",
                                          "Dice_patient"),
                cls_metrics: Sequence[str] = ("cls_AUROC", "cls_Sensitivity",
                                              "cls_Specificity", "cls_F1")
                ) -> pd.DataFrame:
    """Camera-ready main table: metric ± 95 % CI, one row per model."""
    rows: List[Dict[str, object]] = []
    for _, r in summary.iterrows():
        row: Dict[str, object] = {"Model": r.get("display", r["tag"]),
                                  "Params (M)": round(float(r.get("params_M", np.nan)), 1)}
        for m in metrics:
            if m not in r:
                continue
            if ci is not None:
                sub = ci[(ci.tag == r["tag"]) & (ci.metric == m)]
                if len(sub):
                    s = sub.iloc[0]
                    row[m] = format_mean_ci(s.point, s.ci_low, s.ci_high,
                                            2 if m in ("HD95", "ASSD") else 3)
                    continue
            v = float(r[m])
            row[m] = f"{v:.2f}" if m in ("HD95", "ASSD") else f"{v:.3f}"
        for m in cls_metrics:
            if m in r and np.isfinite(float(r[m])):
                row[m.replace("cls_", "")] = f"{float(r[m]):.3f}"
        rows.append(row)
    return pd.DataFrame(rows)


def write_workbook(path: Optional[str] = None, index_first: bool = True) -> str:
    """Write every registered table into one formatted workbook."""
    path = path or os.path.join(OUT_ROOT, "ICH_Hybrid_Results.xlsx")
    tables = registered()
    if not tables:
        raise RuntimeError("no tables registered — run the pipeline first")

    contents = pd.DataFrame(
        [{"sheet": _sheet_name(k), "table": k, "rows": len(v), "columns": len(v.columns),
          "description": _CAPTIONS.get(k, "")} for k, v in tables.items()]
    )

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        if index_first:
            contents.to_excel(xl, sheet_name="00_contents", index=False)
        for name, df in tables.items():
            round_numeric(df).to_excel(xl, sheet_name=_sheet_name(name), index=False)

    _style_workbook(path)
    return path


def _style_workbook(path: str) -> None:
    """Freeze headers, bold them, autofit columns, add an autofilter."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    head_fill = PatternFill("solid", fgColor="D9E5F7")
    head_font = Font(bold=True, color="0D366B")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            width = 10
            for row in range(1, min(ws.max_row, 300) + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    width = max(width, min(len(str(v)) + 2, 46))
            ws.column_dimensions[letter].width = width
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def figure_manifest(captions: Dict[str, str]) -> pd.DataFrame:
    """Which figure file corresponds to which paper figure, with its caption."""
    rows = []
    for i, (stem, cap) in enumerate(captions.items(), start=1):
        png = os.path.join(DIR_FIGURES, f"{stem}.png")
        rows.append({
            "figure_no": f"Fig. {i}",
            "file_png": os.path.relpath(png, OUT_ROOT) if os.path.exists(png) else "MISSING",
            "file_pdf": os.path.relpath(png.replace(".png", ".pdf"), OUT_ROOT)
                        if os.path.exists(png.replace(".png", ".pdf")) else "MISSING",
            "caption": cap,
        })
    return pd.DataFrame(rows)


def write_markdown_summary(path: str, cfg, main_table: pd.DataFrame,
                           tests: Optional[pd.DataFrame], ablation: Optional[pd.DataFrame],
                           captions: Dict[str, str]) -> str:
    """A short, paste-ready results section pointing at every artefact."""
    L: List[str] = []
    L.append("# Results — hybrid CNN–Transformer ICH segmentation and detection\n")
    if len(cfg.folds) > 1:
        protocol = (f"{len(cfg.folds)}-fold patient-wise cross-validation "
                    f"(folds {cfg.folds} of {cfg.n_folds})")
    else:
        protocol = (f"single patient-wise hold-out split "
                    f"(fold {cfg.folds[0]} of a {cfg.n_folds}-way stratified partition: "
                    f"~60 % train / 20 % validation / 20 % test patients, no patient "
                    f"appears in two splits)")
    L.append(f"Run `{cfg.run_name}` · {cfg.img_size}px · {cfg.epochs} epochs · "
             f"{protocol} · seed {cfg.seed}\n")

    L.append("\n## Main comparison\n")
    L.append(main_table.to_markdown(index=False))
    L.append(
        "\n\n**How to read the two Dice columns.** They answer different questions "
        "and differ by a factor of three on this cohort, so quoting one without the "
        "other is misleading.\n\n"
        "- `Dice_agg` pools true/false positives and negatives over *every* test "
        "slice, including the ~88 % with no hemorrhage, and then computes one Dice. "
        "It is dominated by the large bleeds and it punishes false positives on "
        "normal slices. This is the quantity the operating point is tuned for.\n"
        "- `Dice` is the *mean over hemorrhagic slices only*. A slice whose small "
        "bleed is missed entirely contributes 0, so this column is driven by the "
        "many tiny lesions (median 273 px, under 1 % of the image) and is far "
        "lower.\n"
        "- `Dice_patient` pools each patient's slices into one volumetric score and "
        "averages over patients. Patients with no hemorrhage are excluded — their "
        "volumetric Dice is 0/0 — and their false-positive burden is reported "
        "separately as `clean_rate` in Table T04.\n\n"
        "Surface metrics (HD95, ASSD) are in millimetres, converted using each "
        "patient's own in-plane spacing; lower is better.\n")

    if tests is not None and not tests.empty:
        L.append("\n\n## Statistical comparison against the proposed model\n")
        L.append("Wilcoxon signed-rank on paired per-slice metrics, Holm-corrected "
                 "across each metric family. Effect size is the rank-biserial "
                 "correlation.\n")
        keep = ["metric", "model_b", "mean_a", "mean_b", "mean_diff", "ci_low", "ci_high",
                "p_wilcoxon", "p_holm", "signif_label", "rank_biserial"]
        keep = [k for k in keep if k in tests.columns]
        L.append(round_numeric(tests[keep], 4).to_markdown(index=False))

    if ablation is not None and not ablation.empty:
        L.append("\n\n## Ablation\n")
        keep = [c for c in ("label", "Dice", "IoU", "HD95", "cls_AUROC", "params_M")
                if c in ablation.columns]
        L.append(round_numeric(ablation[keep], 4).to_markdown(index=False))

    L.append("\n\n## Figures\n")
    for i, (stem, cap) in enumerate(captions.items(), start=1):
        L.append(f"- **Fig. {i}** — `figures/{stem}.png` — {cap}")

    L.append("\n\n## Files\n")
    L.append("- `ICH_Hybrid_Results.xlsx` — every table as a sheet")
    L.append("- `tables/*.csv` — the same tables as flat CSV")
    L.append("- `figures/*.png|pdf` — all figures at 300 dpi plus vector PDF")
    L.append("- `predictions/*.csv` — per-slice metrics and probabilities "
             "(re-runnable statistics)")
    L.append("- `logs/*.csv` — per-epoch training history")

    text = "\n".join(L) + "\n"
    with open(path, "w") as fh:
        fh.write(text)
    return path
